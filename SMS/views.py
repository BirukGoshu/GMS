from django.shortcuts import render
from rest_framework import viewsets, generics, permissions, serializers, response
from rest_framework.viewsets import GenericViewSet
from rest_framework.mixins import CreateModelMixin, ListModelMixin
import requests, urllib.parse, ssl
from GMS.settings import ozekiusername, ozekipassword, ozekihttpurl
from .serializers import *
from .models import *
from UserManagment.models import *
from csv import DictReader
import openpyxl
# Create your views here.
class SMSViewSet(GenericViewSet, ListModelMixin):
    queryset = SMS.objects.none()
    serializer_class = SMSSerializer
    permission_classes = [permissions.AllowAny]


    def post(self,request):
        Phone = request.data['Recipient']
        Message = request.data['Message']
        Sender=request.data['Sender']
        # messagetype = "SMS:TEXT"
        messagetype='SMS:TEXT:GSM7BIT:CLASS0'
        # messagetype='SMS:TEXT:GSM7BIT:CLASS1'
        recipient = urllib.parse.quote(Phone)
        messagedata = urllib.parse.quote(Message)
        sender = urllib.parse.quote(Sender)
        print(f'sender {sender}')
        # username = ozekiUsername
        # password = ozekiPassword
        # httpUrl = ozekihttpurl
        sendString = (ozekihttpurl + "api?action=sendmessage" + "&username=" + ozekiusername + "&password=" + ozekipassword + "&recipient=" + recipient + "&messagetype=" + messagetype + "&messagedata=" + messagedata + "&responseformat=json"+ "&originator=" + sender)
        requests.packages.urllib3.disable_warnings()
        resp = requests.get(sendString, verify=False)
        # if resp.status_code == 200:
        #     SMS.objects.create(
        #         Recipient = Phone,
        #         Message = messagedata,
        #         # creator = request.user,
        #     )
        return response.Response(resp)

class BulkSMSViewSet(GenericViewSet,CreateModelMixin):
    queryset = BulkSMS.objects.all()
    serializer_class = BulkSMSSerializer
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        file = request.FILES['phones']
        wb = openpyxl.load_workbook(file)
        message = request.POST['message']
        # file.seek(0)
        print(wb['Sheet1'])
        # files = file.file
        # for row in wb.iter_rows():
        #     child=customer_phones(customer_id=request.user, name=row['name'], phone=row['phone'])
        #     child.save()
        excel_data = list()
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
                # child=customer_phones(customer_id=request.user, name=row['name'], phone=row['phone'])
                # child.save()
        print(excel_data)
        phonebook = excel_data
        for phone in phonebook:
            messagetype = "SMS:TEXT"
            # recipient = "+251"
            if phone[1] != 'phone':
                recipient = urllib.parse.quote("+251" + phone[1])
                messagedata = urllib.parse.quote(message)
                sendString = (ozekihttpurl + "api?action=sendmessage" + "&username=" + ozekiusername + "&password=" + ozekipassword + "&recipient=" + recipient + "&messagetype=" + messagetype + "&messagedata=" + messagedata + "&responseformat=json")
                requests.packages.urllib3.disable_warnings()
                resp = requests.get(sendString, verify=False)
        bulk = BulkSMS.objects.create(
            phones = file,
            message = message,
            # creator = request.user,
        )
        return response.Response(resp.text)

class AllCampaignViewSet(viewsets.ModelViewSet):
    queryset = all_campaigns.objects.select_related('customer').select_related('service').order_by('-created_at')
    serializer_class = AllCampaingSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):                                            # added string
        # product=Products.objects.filter(user=self.request.user)
        # transit =
        # for pro in product
        if self.request.user.is_superuser:
            return super().get_queryset().all()
        else:
            return super().get_queryset().filter(customer=self.request.user)
        # return transit

    def update(self, request, pk=None):
        serializer = AllCampaingSerializer(all_campaigns.objects.get(id=request.data.get('id')),data=request.data,context={'request':request}, partial=True)
        # print(serializer)
        if serializer.is_valid():
            serializer.save()
            return response.Response(serializer.data)
        else:
            # print(serializer)
            return response.Response(serializer.errors, status=422)

class OnlyDataCampaignViewSet(viewsets.ModelViewSet):
    queryset = only_data_campaigns.objects.select_related('customer','service').order_by('-created_at')
    serializer_class = OnlyDataCampaignSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):                                            # added string
        # product=Products.objects.filter(user=self.request.user)
        # transit =
        # for pro in product
        if self.request.user.is_superuser:
            transit = super().get_queryset().all()
        else:
            transit = super().get_queryset().filter(customer=self.request.user)
        return transit

    def update(self, request, pk=None):
        serializer = OnlyDataCampaignSerializer(only_data_campaigns.objects.get(id=request.data.get('id')),data=request.data,context={'request':request}, partial=True)
        # print(serializer)
        if serializer.is_valid():
            serializer.save()
            return response.Response(serializer.data)
        else:
            # print(serializer)
            return response.Response(serializer.errors, status=422)

class AllCampaignWhitelistViewSet(viewsets.ModelViewSet):
    queryset=all_campaign_whitelist.objects.select_related('customer','service').order_by('-created_at')
    serializer_class=AllCampaignWhitelistSerialzier
    permission_classes=[permissions.IsAuthenticated]

    def get_queryset(self):                                            # added string
        # product=Products.objects.filter(user=self.request.user)
        # transit =
        # for pro in product
        if self.request.user.is_superuser:
            transit = super().get_queryset().all()
        else:
            transit = super().get_queryset().filter(customer=self.request.user)
        return transit

    def update(self, request, pk=None):
        serializer = AllCampaignWhitelistSerialzier(all_campaign_whitelist.objects.get(id=request.data.get('id')),data=request.data,context={'request':request}, partial=True)
        # print(serializer)
        if serializer.is_valid():
            serializer.save()
            return response.Response(serializer.data)
        else:
            # print(serializer)
            return response.Response(serializer.errors, status=422)

class SurveyViewSet(viewsets.ModelViewSet):
    queryset=ordinarysurvey.objects.select_related('customer','service').order_by('-created_at')
    serializer_class=SurveySerializer
    permission_classes=[permissions.IsAuthenticated]

    def get_queryset(self):                                            # added string
        # product=Products.objects.filter(user=self.request.user)
        # transit =
        # for pro in product
        if not self.request.user.is_superuser:
            transit = super().get_queryset().all()
        else:
            transit = super().get_queryset().filter(customer=self.request.user)
        return transit

    def update(self, request, pk=None):
        serializer = SurveySerializer(survey.objects.get(id=request.data.get('id')),data=request.data,context={'request':request}, partial=True)
        # print(serializer)
        if serializer.is_valid():
            serializer.save()
            return response.Response(serializer.data)
        else:
            # print(serializer)
            return response.Response(serializer.errors, status=422)

class DemographicSurveyViewSet(viewsets.ModelViewSet):
    queryset=demographicsurvey.objects.select_related('customer','service').order_by('-created_at')
    serializer_class=DemographicSurveySerializer
    permission_classes=[permissions.IsAuthenticated]

    def get_queryset(self):                                            # added string
        # product=Products.objects.filter(user=self.request.user)
        # transit =
        # for pro in product
        if not self.request.user.is_superuser:
            transit = super().get_queryset().all()
        else:
            transit = super().get_queryset().filter(customer=self.request.user)
        return transit

class InternalSurveyViewSet(viewsets.ModelViewSet):
    queryset=internalsurvey.objects.select_related('customer','service').order_by('-created_at')
    serializer_class=InternalSurveySerializer
    permission_classes=[permissions.IsAuthenticated]

    def get_queryset(self):                                            # added string
        # product=Products.objects.filter(user=self.request.user)
        # transit =
        # for pro in product
        if not self.request.user.is_superuser:
            transit = super().get_queryset().all()
        else:
            transit = super().get_queryset().filter(customer=self.request.user)
        return transit

class BussinessTypeViewSet(viewsets.ModelViewSet):
    queryset = bussiness_types.objects.all()
    serializer_class = BussinessTypeSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

class TemplateViewSet(viewsets.ModelViewSet):
    queryset=template.objects.all()
    serializer_class=TemplateSerializer
    permission_classes=[permissions.IsAuthenticated]

class IndustryViewSet(viewsets.ModelViewSet):
    queryset = industries.objects.all()
    serializer_class = IndustrySerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

class PackageSubscriptionViewSet(viewsets.ModelViewSet):
    queryset = package_subscriptions.objects.all()
    serializer_class = PackageSubscriptionSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

class CategoryViewSet(viewsets.ModelViewSet):
    queryset=Category.objects.all()
    serializer_class=CategorySerialzier
    permission_classes = [permissions.IsAuthenticated]

class SubCategoryViewSet(viewsets.ModelViewSet):
    queryset=SubCategories.objects.all()
    serializer_class=SubCategorySerializer
    permission_classes = [permissions.IsAuthenticated]

class CustomCategoryViewSet(viewsets.ModelViewSet):
    queryset=CustomCategory.objects.all()
    serializer_class=CustomCategorySerializer
    permission_classes = [permissions.IsAuthenticated]

class ServiceViewSet(viewsets.ModelViewSet):
    queryset = services.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

class WalletViewSet(viewsets.ModelViewSet):
    # id = users.objects.all()
    queryset = wallet.objects.all()
    serializer_class = WalletSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.is_superuser:
            return super().get_queryset()
        else:
            return super().get_queryset().filter(user=self.request.user.id)

class UserWalletViewSet(viewsets.ModelViewSet):
    queryset = userwallet.objects.all()
    serializer_class = UserWalletSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.is_superuser:
            return super().get_queryset()
        else:
            return super().get_queryset().filter(user=self.request.user.id)

class WalletPaymentViewSet(GenericViewSet,CreateModelMixin):
    model=wallet
    serializer_class=paymentserializer
    permission_classes=[permissions.IsAuthenticated]

def verifywalletpayment(request, id,slug):
    chapa = Chapa(chapakey, response_format='json')
    resp = chapa.verify(slug)
    w=wallet.objects.get(id=id)
    wallets=userwallet.objects.get(user=w.user)
    wallets.balance += int(resp['data']['amount'])
    w.paymentdata=resp
    w.save()
    wallets.save()
    # insurance.paymentdata=resp
    # if resp['status']=='success':
    #     insurance.status='waiting for payment confirmation'
    # else:
    #     insurance.status='payment failed'
    # insurance.save()


