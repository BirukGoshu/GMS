from django.contrib import admin
from .models import *
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.utils import timezone
from csv import DictReader
import openpyxl
import requests, urllib.parse, ssl
from GMS.settings import ozekiusername, ozekipassword, ozekihttpurl
from rest_framework import response
from UserManagment.models import customer_phones, subscribers
from itertools import chain
# Register your models here.
admin.site.register(all_campaigns)
admin.site.register(only_data_campaigns)
admin.site.register(wallet)
admin.site.register(template)
admin.site.register(all_campaign_whitelist)
admin.site.register(ordinarysurvey)
admin.site.register(demographicsurvey)
admin.site.register(internalsurvey)

@receiver(post_save, sender=all_campaigns)
def ApproveCampaign(instance, sender, *args, **kwargs):
    # print('approving')
    # print(instance.approved)
    if instance.phones is not None:
        file = instance.phones
        wb = openpyxl.load_workbook(file)
        message = instance.message
        excel_data = list()
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
        print(excel_data)
        phonebook = excel_data
        totalphones=0
        for phone in phonebook:
            # if not isinstance(phone[0],str):
            if phone[0] != 'phone' and phone[0] != 'None':
                totalphones += 1
            if phone[0][0] == '9' or phone[0][0] =='7':
                recipient = "+251" + phone[0]
                if not customer_phones.objects.filter(phone=recipient,customer=instance.customer):
                    customer_phones.objects.create(
                        customer = instance.customer,
                        phone = recipient
                    )
                    print(f'{phone[0]} added to customer phone \n initial {phone[0][0]}')
            elif phone[0][0]=='2':
                recipient='+'+phone[0]
                if not customer_phones.objects.filter(phone=recipient):
                    customer_phones.objects.create(
                        customer = instance.customer,
                        phone = recipient
                    )
                    print(f'{phone[0]} added to customer phone \n initial {phone[0][0]}')
            else:
                print(f'{phone[0]} not valid format in all_data_campaign ')
        if totalphones !=0:
            instance.totalphones=totalphones

    if instance.approved==False and instance.sent == False:
        # file = instance.phones
        # wb = openpyxl.load_workbook(file)
        # message = instance.message
        # print('approved')
        # # file.seek(0)
        # print(wb['Sheet1'])
        # # files = file.file
        # # for row in wb.iter_rows():
        # #     child=customer_phones(customer_id=request.user, name=row['name'], phone=row['phone'])
        # #     child.save()
        # excel_data = list()
        # for name in wb.sheetnames:
        #     sheet = wb[name]
        #     for row in sheet.iter_rows():
        #         row_data = list()
        #         for cell in row:
        #             row_data.append(str(cell.value))
        #         excel_data.append(row_data)
        #         # child=customer_phones(customer_id=request.user, name=row['name'], phone=row['phone'])
        #         # child.save()
        # print(excel_data)
        # phonebook = excel_data
        # if 'messagetype' in self.context['request'].data:
        service=instance.service
        if service:
            if 'flash' in service.name or 'smart' in service.name:
                messagetype='SMS:TEXT:GSM7BIT:CLASS0'
            elif 'promotion' in service.name:
                messagetype='SMS:TEXT:GSM7BIT:CLASS1'
            else:
                messagetype='SMS:TEXT'
        else:
            messagetype='SMS:TEXT'
        # messagetype = "SMS:TEXT"
        originator = urllib.parse.quote(instance.sender)
        recipientlist = ""
        i = 0
        messagedata = urllib.parse.quote(message)
        # date = urllib.parse.quote(instance.date_tobe_sent)
        date = str(instance.date_tobe_sent.date()) + '%20' + str(instance.date_tobe_sent.time())
        print(date)
        for phone in phonebook:
            if phone[0] != 'phone':
                if phone[0][0] == '9' or phone[0][0] =='7':
                    recipient = urllib.parse.quote("+251" + phone[0])
                elif phone[0][0]=='2':
                    recipient=urllib.parse.quote('+'+phone[0])
                else:
                    print(f'{phone[0]} not valid format in all_data_campaign from {instance.customer.username}')
                recipientlist +=  "&recepient{}=".format(i) + recipient + "&messagedata{}=".format(i) + messagedata + "&messagetype{}=".format(i) + messagetype + "&originator{}=".format(i) + originator + "&sendondate{}=".format(i) + date
                i += 1
        # w = wallet.objects.get(user=instance.customer)
        # if w.balance >= 0.3 * (len(phonebook)-1):
        sendString = (ozekihttpurl + "api?action=sendmessage" + "&messagecount=" + str(len(phonebook)-1) + "&username=" + ozekiusername + "&password=" + ozekipassword + recipientlist  + "&responseformat=json"  )
        requests.packages.urllib3.disable_warnings()
        resp = requests.get(sendString, verify=False)
            # if resp.status_code == 200:
                # w.balance -= 0.3 * (len(phonebook)-1)
                # w.save()
        # else:
            # return response.Response('Insufficient Balance')
        print(recipientlist)
        print(sendString)
        instance.sent = True
        instance.save()

@receiver(post_save, sender=only_data_campaigns)
def ApproveOnlyDataCampaign(instance, sender, *args, **kwargs):
    if instance.approved != True and instance.sent != True:
        # phonebook = only_data_campaigns.objects.none()
        phonebook=[]
        if instance.industry != None:
            phonebook += dict(chain(phonebook,subscribers.objects.filter(industry=instance.industry) ))
        if instance.minage is not None and instance.maxage is not None:
            phonebook +=dict(chain(phonebook,subscribers.objects.filter(age__range=(instance.minage,instance.maxage))))
            # for age in instance.age:
            #     phonebook += subscribers.objects.filter(age=age)
        if instance.phone_type != None:
            phonebook += dict(chain(phonebook,subscribers.objects.filter(phonetype=instance.phone_type)))
        service=instance.service
        if service:
            if 'flash' in service.name:
                messagetype='SMS:TEXT:GSM7BIT:CLASS0'
            elif 'promotion' in service.name:
                messagetype='SMS:TEXT:GSM7BIT:CLASS1'
            else:
                messagetype='SMS:TEXT'
        else:
            messagetype='SMS:TEXT'
        # messagetype = "SMS:TEXT"
        originator = urllib.parse.quote(instance.campaign_name)
        recipientlist = ""
        i = 0
        messagedata = urllib.parse.quote(instance.message)
        # date = urllib.parse.quote(instance.date_tobe_sent)
        date = str(instance.date_tobe_sent.date()) + '%20' + str(instance.date_tobe_sent.time())
        print(date)
        totalphones=0
        for phone in phonebook:
            if phone.phone != 'phone' and phone.phone !='None':
                totalphones +=1
            recipient = urllib.parse.quote(phone.phone)
            recipientlist +=  "&recepient{}=".format(i) + recipient + "&messagedata{}=".format(i) + messagedata + "&messagetype{}=".format(i) + messagetype + "&originator{}=".format(i) + originator + "&sendondate{}=".format(i) + date
            i += 1
        # + "&messagetype=" + messagetype + "&messagedata=" + messagedata
        # w = wallet.objects.get(user=instance.customer)
        # if w.balance >= 0.3 * (len(phonebook)-1):
        sendString = (ozekihttpurl + "api?action=sendmessage" + "&messagecount=" + str(len(phonebook)-1) + "&username=" + ozekiusername + "&password=" + ozekipassword + recipientlist  + "&responseformat=json"  )
        requests.packages.urllib3.disable_warnings()
        resp = requests.get(sendString, verify=False)
        if totalphones != 0:
            instance.totalphones = totalphones
        #     if resp.status_code == 200:
        #         w.balance -= 0.3 * (len(phonebook)-1)
        #         w.save()
        # else:
        #     return response.Response('Insufficient Balance')
        print(recipientlist)
        print(sendString)
        instance.sent = True
        instance.save()